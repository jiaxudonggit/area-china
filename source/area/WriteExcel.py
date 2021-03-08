# -*- coding: UTF-8 -*-
"""
@description 写入excel
@author jxd
"""

import os
from openpyxl import Workbook
from concurrent.futures import ThreadPoolExecutor,  wait, ALL_COMPLETED


class WriteExcel(object):

    def __init__(self, file_name: str, data: dict):
        """
        :param file_name: 文件名
        :param data: 要写入的数据
        """
        # 文件路径
        self.ROOT = os.path.join(os.path.abspath(os.path.dirname(__file__)), '/result')
        # 表头
        self.HEADER = [
            "统计用区划代码",
            "名称",
            "统计用区划代码",
            "名称",
            "统计用区划代码",
            "名称",
            "统计用区划代码",
            "城乡分类代码",
            "名称",
        ]
        self.file_name = file_name
        self.data = data
        self.path = os.path.join(self.ROOT, self.file_name)
        self.work_book = Workbook()

    def write(self, item1):

        # 创建sheet
        work_sheet_total = self.work_book.create_sheet(title=item1.get("name"))
        # 插入表头
        work_sheet_total.append(self.HEADER)
        # 一级
        if not item1.get("child"):
            return False

        for code2, item2 in item1.get("child").items():
            # 二级
            data2 = []
            data2.append(item2.get("code"))
            data2.append(item2.get("name"))
            if not item2.get("child"):
                # 插入数据
                work_sheet_total.append(data2)
                continue
            for code3, item3 in item1.items():
                # 三级
                data3 = []
                data3.append(item2.get("code"))
                data3.append(item2.get("code"))
                data3.append(item3.get("code"))
                data3.append(item3.get("name"))
                if not item3.get("child"):
                    # 插入数据
                    work_sheet_total.append(data3)
                    continue
                for code4, item4 in item1.items():
                    data4 = []
                    # 三级
                    data4.append(item2.get("code"))
                    data4.append(item2.get("code"))
                    data4.append(item3.get("code"))
                    data4.append(item3.get("name"))
                    data4.append(item4.get("code"))
                    data4.append(item4.get("name"))
                    if not item4.get("child"):
                        # 插入数据
                        work_sheet_total.append(data4)
                        continue
                    for code5, item5 in item1.items():
                        data5 = []
                        # 三级
                        data5.append(item2.get("code"))
                        data5.append(item2.get("code"))
                        data5.append(item3.get("code"))
                        data5.append(item3.get("name"))
                        data5.append(item4.get("code"))
                        data5.append(item4.get("name"))
                        data5.append(item5.get("code"))
                        data5.append(item5.get("code_type"))
                        data5.append(item5.get("name"))

                        work_sheet_total.append(data5)

        return True

    def write_one_thread(self):
        """
        单线程写入
        :return:
        """
        for code1, item1 in self.data.items():
            # 创建sheet
            work_sheet_total = self.work_book.create_sheet(title=item1.get("name"))
            # 插入表头
            work_sheet_total.append(self.HEADER)
            # 一级
            if not item1.get("child"):
                continue
            for code2, item2 in item1.get("child").items():
                # 二级
                data2 = []
                data2.append(item2.get("code"))
                data2.append(item2.get("name"))
                if not item2.get("child"):
                    # 插入数据
                    work_sheet_total.append(data2)
                    continue
                for code3, item3 in item1.items():
                    # 三级
                    data3 = []
                    data3.append(item2.get("code"))
                    data3.append(item2.get("code"))
                    data3.append(item3.get("code"))
                    data3.append(item3.get("name"))
                    if not item3.get("child"):
                        # 插入数据
                        work_sheet_total.append(data3)
                        continue
                    for code4, item4 in item1.items():
                        data4 = []
                        # 三级
                        data4.append(item2.get("code"))
                        data4.append(item2.get("code"))
                        data4.append(item3.get("code"))
                        data4.append(item3.get("name"))
                        data4.append(item3.get("code"))
                        data4.append(item3.get("name"))
                        if not item4.get("child"):
                            # 插入数据
                            work_sheet_total.append(data4)
                            continue
                        for code5, item5 in item1.items():
                            data5 = []
                            # 三级
                            data5.append(item2.get("code"))
                            data5.append(item2.get("code"))
                            data5.append(item3.get("code"))
                            data5.append(item3.get("name"))
                            data5.append(item3.get("code"))
                            data5.append(item3.get("name"))
                            data5.append(item3.get("code"))
                            data5.append(item3.get("code_type"))
                            data5.append(item3.get("name"))

                            work_sheet_total.append(data5)

        if not os.path.exists(self.path):
            os.makedirs(self.ROOT)

        # 保存到本地
        self.work_book.save(self.path)

    def write_multi_thread(self):
        """
        多线程写入，只写入一个一级行政区时，和单线程没区别
        :return:
        """
        with ThreadPoolExecutor(max_workers=6) as t:  # 创建一个最大容纳数量为6的线程池
            all_task = []
            for code, item in self.data.items():
                task = t.submit(self.write, item)
                all_task.append(task)

            wait(all_task, return_when=ALL_COMPLETED)

            if not os.path.exists(self.path):
                os.makedirs(self.ROOT)

            # 保存到本地
            self.work_book.save(self.path)